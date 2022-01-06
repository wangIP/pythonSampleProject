import tkinter as tk
from tkinter.constants import S
import module.clowerWeb as m
import openpyxl
from openpyxl.styles import Alignment
def clickSearch():
    itemListName=[]
    itemListText=[]
    itemListNameNew=[]
    itemListTextNew=[]
    status=False
    keyName = keyWorld.get()
    
    itemListName,itemListText=m.click()
    for x in range(len(itemListName)) :
        if keyName in itemListName[x] or keyName in itemListText[x]:
            # print(itemList[x])
            # itemListNew.append(itemList[x])
            itemListNameNew.append(itemListName[x])
            itemListTextNew.append(itemListText[x])
            status=True
    if status == True:
        fill = openpyxl.styles.PatternFill(patternType='solid',
                                   fgColor='FF0000', bgColor='FF0000')
        wb=openpyxl.Workbook()
        s1 = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        s1.cell(row=1,column=1,value="Name").fill=fill
        s1.cell(row=1,column=7,value="Text").fill=fill
        t=0
        for i in range(0,len(itemListNameNew)):
            # A列
                s1.cell(row=t+2,column=1,value=itemListNameNew[i])
            # B列
                s1.cell(row=t+2,column=7,value=itemListTextNew[i])
                s1.merge_cells('G'+str(t+2)+':M'+str(t+7))
                # 結合したセルには範囲の左上のセルのみに設定する。
                s1['G'+str(t+2)].alignment=Alignment(horizontal='justify')
                t +=7           
        wb.save('test.xlsx')
    # print(itemList)

windows = tk.Tk()
windows.geometry('640x480')
windows.title('Font Download')

inputFm = tk.Frame(windows,bg='black',width=640,height=120)
inputFm.pack()

lb = tk.Label(inputFm,text='please enter you want search keyworld',bg='red',fg='white',font=('細明體',12))
lb.place(rely=0.25,relx=0.5,anchor='center')

keyWorld = tk.StringVar()
entry=tk.Entry(inputFm,textvariable=keyWorld,width=50)
entry.place(rely=0.5,relx=0.5,anchor='center')

inputButton = tk.Button(inputFm,text='search',command=clickSearch,bg='red',fg='Black',font=('細明體',10))
inputButton.place(rely=0.5,relx=0.85,anchor='center')

windows.mainloop()